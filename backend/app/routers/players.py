import io
import uuid

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import get_current_user, is_module_enabled_for_club, require_module
from app.models import Player, PlayerWeeklyDistanceLog, User
from app.models import TrainingCycle as DbTrainingCycle
from app.routers._readiness import _normalized_wellness, load_squad_readiness
from app.schemas import (
    PlayerCreate,
    PlayerImportError,
    PlayerImportResult,
    PlayerOut,
    PlayerUpdate,
    SquadOverviewPlayerSchema,
    WeeklyDistanceOut,
)
from app.services.platform_admin import ModuleKey

router = APIRouter(
    prefix="/api/players", tags=["players"], dependencies=[Depends(require_module(ModuleKey.SQUAD_OVERVIEW))]
)

TEMPLATE_HEADERS = ["Rugnummer", "Naam", "Voornaam", "E-mailadres", "Telefoonnummer"]


@router.get("/import-template")
def download_import_template(current_user: User = Depends(get_current_user)):
    """Downloadable .xlsx template coaches fill in and re-upload via POST /players/import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spelers"
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append([7, "Peeters", "Jan", "jan.peeters@example.be", "0470 12 34 56"])
    for cell in ws[2]:
        cell.font = Font(italic=True, color="999999")
    for col, width in zip("ABCDE", [12, 20, 20, 30, 20]):
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="spelers_import_sjabloon.xlsx"'},
    )


@router.post("/import", response_model=PlayerImportResult)
def import_players(
    file: UploadFile,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload een .xlsx-bestand (gebruik het downloadbare sjabloon)")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Kon het Excel-bestand niet lezen") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Bestand is leeg")

    created = 0
    skipped = 0
    errors: list[PlayerImportError] = []

    for row_number, row in enumerate(rows[1:], start=2):  # skip header row
        if row is None or all(cell in (None, "") for cell in row):
            continue  # silently skip fully blank rows

        cells = list(row) + [None] * (5 - len(row))
        rugnummer, naam, voornaam, email, telefoonnummer = cells[:5]

        if not naam or not voornaam:
            skipped += 1
            errors.append(PlayerImportError(row=row_number, message="Naam en voornaam zijn verplicht"))
            continue

        jersey_number = None
        if rugnummer not in (None, ""):
            try:
                jersey_number = int(rugnummer)
            except (TypeError, ValueError):
                skipped += 1
                errors.append(PlayerImportError(row=row_number, message=f"Ongeldig rugnummer: {rugnummer!r}"))
                continue

        player = Player(
            club_id=current_user.club_id,
            first_name=str(voornaam).strip(),
            last_name=str(naam).strip(),
            email=str(email).strip() if email not in (None, "") else None,
            phone_number=str(telefoonnummer).strip() if telefoonnummer not in (None, "") else None,
            jersey_number=jersey_number,
        )
        db.add(player)
        created += 1

    db.commit()
    return PlayerImportResult(created=created, skipped=skipped, errors=errors)


@router.post("", response_model=PlayerOut, status_code=201)
def create_player(
    payload: PlayerCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    player = Player(club_id=current_user.club_id, **payload.model_dump())
    db.add(player)
    db.commit()
    db.refresh(player)
    return player


@router.get("", response_model=list[PlayerOut])
def list_players(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.scalars(
        select(Player).where(Player.club_id == current_user.club_id).order_by(Player.jersey_number.nulls_last())
    ).all()


@router.get("/squad-overview", response_model=list[SquadOverviewPlayerSchema])
def squad_overview(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Per-player readiness status for the Squad Overview grid: classifies
    each player as fit / reductie / overbelast, reusing the exact
    flag_players() rules the Next Training proposal uses
    (app.services.team_readiness) so the two stay consistent. Km-based
    (acute/chronic distance) is the always-available basis; RPE/wellness
    only adds to it when ModuleKey.RPE_WELLNESS is active for this club.
    'geen_data' applies only when a player has neither real km data
    (chronic_km_28d == 0) nor an RPE entry — not just for lacking one of
    the two.

    flag_players() is keyed by player_name (not id) — an existing
    limitation of that service (two same-named players in one club would
    collide), not something introduced here."""
    rpe_module_active = is_module_enabled_for_club(current_user.club_id, ModuleKey.RPE_WELLNESS, db)
    squad = load_squad_readiness(current_user.club_id, db, rpe_module_active=rpe_module_active)

    result = []
    for sr in squad:
        player = sr.player
        readiness = sr.readiness
        latest = sr.latest

        if readiness.chronic_km_28d <= 0 and latest is None:
            result.append(
                SquadOverviewPlayerSchema(
                    id=player.id,
                    first_name=player.first_name,
                    last_name=player.last_name,
                    jersey_number=player.jersey_number,
                    position=player.position,
                    status="geen_data",
                    acwr=None,
                    latest_rpe=None,
                    latest_wellness=None,
                    flags=[],
                )
            )
            continue

        flag_types = {f.flag_type for f in sr.flags}
        if "injured" in flag_types or "overload" in flag_types:
            status = "overbelast"
        elif "poor_recovery" in flag_types or "underload" in flag_types or "acwr_trending_up" in flag_types:
            status = "reductie"
        else:
            status = "fit"

        result.append(
            SquadOverviewPlayerSchema(
                id=player.id,
                first_name=player.first_name,
                last_name=player.last_name,
                jersey_number=player.jersey_number,
                position=player.position,
                status=status,
                acwr=(
                    round(readiness.acute_km_7d / readiness.chronic_km_28d, 2)
                    if readiness.chronic_km_28d > 0
                    else None
                ),
                acwr_rpe=(
                    round(readiness.acute_load_7d / readiness.chronic_load_28d, 2)
                    if readiness.acute_load_7d is not None
                    and readiness.chronic_load_28d is not None
                    and readiness.chronic_load_28d > 0
                    else None
                ),
                latest_rpe=latest.rpe_score if latest else None,
                latest_wellness=_normalized_wellness(latest) if latest else None,
                flags=[f.detail for f in sr.flags],
                flag_types=[f.flag_type for f in sr.flags],
                flag_sources=[f.source for f in sr.flags],
            )
        )
    return result


@router.get("/{player_id}", response_model=PlayerOut)
def get_player(player_id: uuid.UUID, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    player = db.get(Player, player_id)
    if player is None or player.club_id != current_user.club_id:
        raise HTTPException(status_code=404, detail="Player not found")
    return player


@router.get("/{player_id}/weekly-distance", response_model=WeeklyDistanceOut)
def get_weekly_distance(
    player_id: uuid.UUID,
    week_number: int = Query(ge=1),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Summed match_distance_km (auto-logged when match minutes are saved, see
    app/routers/matches.py) + training_distance_km (reserved for a future
    completed-training log; always 0 for now) for one player in one week of
    the club's currently active cycle — for showing next to a position's km
    target on the frontend."""
    player = db.get(Player, player_id)
    if player is None or player.club_id != current_user.club_id:
        raise HTTPException(status_code=404, detail="Player not found")

    active_cycle = db.scalar(
        select(DbTrainingCycle).where(
            DbTrainingCycle.club_id == current_user.club_id, DbTrainingCycle.is_active.is_(True)
        )
    )
    if active_cycle is None:
        return WeeklyDistanceOut(
            player_id=player_id, week_number=week_number, match_distance_km=0.0,
            training_distance_km=0.0, total_km=0.0,
        )

    totals = db.execute(
        select(
            func.coalesce(func.sum(PlayerWeeklyDistanceLog.match_distance_km), 0),
            func.coalesce(func.sum(PlayerWeeklyDistanceLog.training_distance_km), 0),
        ).where(
            PlayerWeeklyDistanceLog.player_id == player_id,
            PlayerWeeklyDistanceLog.training_cycle_id == active_cycle.id,
            PlayerWeeklyDistanceLog.week_number == week_number,
        )
    ).one()
    match_distance_km, training_distance_km = float(totals[0]), float(totals[1])

    return WeeklyDistanceOut(
        player_id=player_id,
        week_number=week_number,
        match_distance_km=match_distance_km,
        training_distance_km=training_distance_km,
        total_km=round(match_distance_km + training_distance_km, 2),
    )


@router.patch("/{player_id}", response_model=PlayerOut)
def update_player(
    player_id: uuid.UUID,
    payload: PlayerUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    player = db.get(Player, player_id)
    if player is None or player.club_id != current_user.club_id:
        raise HTTPException(status_code=404, detail="Player not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(player, field, value)
    db.commit()
    db.refresh(player)
    return player


@router.delete("/{player_id}", status_code=204)
def delete_player(
    player_id: uuid.UUID, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    player = db.get(Player, player_id)
    if player is None or player.club_id != current_user.club_id:
        raise HTTPException(status_code=404, detail="Player not found")
    db.delete(player)
    db.commit()
